import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import datetime
from os import path
import config

#ИЗМЕНЕНИЕ ВРЕМЕНИ ПО ЧАС ПОЯСУ
def get_date_time(ts, timezone, dt_format='%H:%M:%S'):
    tz = datetime.timezone(datetime.timedelta(seconds=timezone))
    return datetime.datetime.fromtimestamp(ts, tz=tz).strftime(dt_format)

#ОПРЕДЕЛЕНИЕ ГОРОДА И ЕГО ПАРАММЕТРОВ
def get_weather(city_name):
    params = {
        'appid': config.API_KEY,
        'units': config.UNITS,
        'lang': config.LANG,
        'q': city_name
    }
    try:
        r = requests.get(config.API_URL, params=params)
        return r.json()
    except:
        return {'cod': 0, 'message': 'Не удалось получить данные'}

#ВЫГРУЗКА ИНФОРМАЦИИ ПО ПОГОДЕ ГОРОДА
def print_weather(data):
    if data['cod'] != 200:
        print(data['message'])
        return {}
    else:
        sunrise_time = get_date_time(data['sys']['sunrise'], data['timezone'])
        sunset_time = get_date_time(data['sys']['sunset'], data['timezone'])
        print(f'''Местоположение: {data['name']}, {data['sys']['country']}
Температура: {data['main']["temp"]} C
Атм. давление: {data['main']['pressure']} rlla
Влажность: {data['main']['humidity']}%
Скорость ветра: {data['wind']['speed']} м/с
Погодные условия: {data['weather'][0]['description']}
Восход: {sunrise_time}
Закат: {sunset_time}
''')
        print('+' * 50)
        return data

#СОХРАНЯЕМ ЗАПРОС В БАЗУ ДАННЫХ
def save_excel(data):
    if data['cod'] == 200:
        if path.exists(config.FILE_EXCEL):
            wb = load_workbook(filename=config.FILE_EXCEL)
            ws = wb.active

        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Статистика запросов'
            ws.append(['Дата запроса', 'Город', 'Температура'])
            ft = Font(color='696969', bold=True)
            a1 = ws['A1']
            b1 = ws['B1']
            c1 = ws['C1']
            a1.font = ft
            b1.font = ft
            c1.font = ft

        ws.append([datetime.datetime.now(), f"{data['name']}, {data['sys']['country']}", data["main"]["temp"]])
        wb.save(filename=config.FILE_EXCEL)



